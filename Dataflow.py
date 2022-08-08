from Query import Query
from Target import Target
from TransformCall import TransformCall
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

class Dataflow:
    def print_nice(self,  *args):
        print(self.indent * '\t', *args)

    def __init__(self, node):
        self.node = node
        self.name = node.attrib['name']
        self.transforms = node.find('DITransforms')
        self.queries = self.transforms.findall('DIQuery')
        self.transform_calls = self.transforms.findall('DITransformCall')
        self.targets = self.transforms.findall('DIDatabaseTableTarget')
        self.indent = 1
        self.p_queries = []
        self.p_transforms = []
        self.p_targets = {}

    def process_dataflows(self):
        self.print_nice('Processing dataflow', self.name)
        self.print_nice('Dataflow', self.name, 'has', len(self.transform_calls), 'transform_calls')
        self.indent += 1
        for call in self.transform_calls:
            self.process_transform_calls(call)
        self.indent -= 1
        self.print_nice('Dataflow', self.name, 'has', len(self.queries), 'queries')
        self.indent += 1
        for query in self.queries:
            self.process_queries(query)
        self.indent -= 1
        self.print_nice('Dataflow', self.name, 'has', len(self.targets), 'targets')
        self.indent += 1
        for target in self.targets:
            self.process_targets(target)
        self.indent -= 1

    def process_queries(self, query):
        query = Query(query)
        self.p_queries.append(query)
        self.print_nice('Found query named', query.name)
        self.print_nice('Query', query.name, 'has', len(query.elements), 'elements')
        self.print_nice('Query', query.name, 'has', len(query.expressions), 'expr')
        query.process(self.name)

    def process_targets(self, target):
        target = Target(target)
        self.p_targets[target.input_view] = self.name
        self.print_nice('Has target named', target.name)

    def process_transform_calls(self, transform_call):
        transform_call = TransformCall(transform_call)
        self.p_transforms.append(transform_call)
        self.print_nice('Has call named', transform_call.name)
        if transform_call.sql_call:
            transform_call.write_sql_call(self.name)

    def create_excel(self):
        #Starting son las coordenadas del primer header
        starting_row = 2
        starting_column = 2
        header = ['DATAFLOW', 'FUENTE', 'TRANSFORMACION', 'CAMPOS', 'CRITERIOS', 'TARGET']
        wb = Workbook()
        ws = wb.active
        for pos in range(len(header)):
            ws.cell(row=starting_row, column=starting_column+pos).value = header[pos]
        ws.cell(row=starting_row+1, column=starting_column).value=self.name
        offset_row = 0
        offset_column = 0
        for query in self.p_queries:
            ws.cell(row=starting_row+1+offset_row, column=starting_column+2).value = query.name
            ws.cell(row=starting_row+1+offset_row, column=starting_column+1).value = query.source
            for key, value in query.transforms.items():
                ws.cell(row=starting_row+1+offset_row, column=starting_column+3+offset_column).value = key
                ws.cell(row=starting_row+1+offset_row, column=starting_column+4+offset_column).value = value
                offset_row+=1
            temp1 = offset_row-len(query.transforms)
            temp2 = offset_row
            print('About to merge', temp1, temp2)
            ws.merge_cells(start_row=starting_row+1+temp1, end_row=starting_row+temp2, start_column=starting_column+1, end_column=starting_column+1)
            ws.merge_cells(start_row=starting_row+1+temp1, end_row=starting_row+temp2, start_column=starting_column+2, end_column=starting_column+2)
            ws.merge_cells(start_row=starting_row+1+temp1, end_row=starting_row+temp2, start_column=starting_column+5, end_column=starting_column+5)

        ws.merge_cells(start_row=starting_row + 1, end_row=starting_row + offset_row,
                       start_column=starting_column, end_column=starting_column)
        wb.save(f'{self.name}.xlsx')